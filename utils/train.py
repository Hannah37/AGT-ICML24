from utils.utility import *
from utils.metric import *
from utils.approximate import *

import wandb
import os
import glob
import openpyxl
import pandas as pd
import torch.optim as optim
import torch.nn.functional as F

### Trainer for AGT
class AGT_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, adj_size, cv_idx, pretrained_net, pretrained_t):
        self.args = args
        self.device = device
        self.network = network.to(device)
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = valid_loader
        self.adj_size = adj_size # num_ROIs
        self.optimizer = optimizer
        self.p_net = pretrained_net.to(self.device)
        self.p_t = pretrained_t.to(self.device)
        self.cv_idx = cv_idx
        
    ### Train
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list = [] # List of val accuracy
        val_sens_list, val_prec_list = [], []

        
        best_val_acc = 0.6
        for epoch in range(1, self.args.epochs + 1):
            self.network.train()
            loss_train_avg, acc_train_avg = [], [] 

            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero

                ### Use heat kernel instead of adjacency matrix
                output, dist_loss = self.network.forward(feature, adjacency, eigenvalue, eigenvector, label) # Shape: (# of samples, # of labels)

                loss_train = self.args.beta * dist_loss + F.nll_loss(output, label) 
                accuracy_train = compute_accuracy(output, label)

                loss_train_avg.append(loss_train.item())
                acc_train_avg.append(accuracy_train.item())

                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves

                self.optimizer.step() # Updates the parameters

                            
                wandb.log({"loss_train": loss_train.item(),
                          "acc_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            print(f"\n Epoch [{epoch} / {self.args.epochs}] loss_train: {sum(loss_train_avg)/len(loss_train_avg):.5f} acc_train: {sum(acc_train_avg)/len(acc_train_avg):.5f}", end="")
            
            # inference
            self.network.eval()
            with torch.no_grad():
                for adjacency, feature, label, eigenvalue, eigenvector in self.valid_loader:
                    
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.p_t) # Use heat kernel instead of adjacency matrix
                    pseudo_label = self.p_net.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                    one_hot_label = torch.argmax(pseudo_label, dim=1).squeeze() 
                    output, dist_loss = self.network.forward(feature, adjacency, eigenvalue, eigenvector, one_hot_label, is_train=False) 

                    loss_val = self.args.beta * dist_loss + F.nll_loss(output, label) 
                    accuracy_val = compute_accuracy(output, label)

                    val_acc_list.append(accuracy_val.item())
                    ac, pr, sp, se, f1 = confusion(output, label)

                    val_sens_list.append(se)
                    val_prec_list.append(pr)

                    print(' loss_val: {:.4f}'.format(loss_val.item()),
                        'acc_val: {:.4f}'.format(accuracy_val.item()), end='')
                    
                    wandb.log({"loss_val": loss_val.item(),
                        "acc_val": accuracy_val.item()
                        })
                    
                    if accuracy_val > best_val_acc:
                        best_val_acc = accuracy_val
                        
                        files = glob.glob(self.args.save_dir + '*.pth')
                        for file in files:
                            os.remove(file) # remove previous saved models

                        torch.save(self.network.state_dict(), os.path.join(self.args.save_dir, '{}.pth'.format(self.cv_idx)))
                        print(' Saved !! ')

        return val_acc_list, val_sens_list, val_prec_list
        
    ### Test
    def load_and_test(self, saved_model, model_path):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        saved_model = saved_model.to(self.device)
        saved_model.load_state_dict(torch.load(model_path))

        saved_model.eval()
        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.p_t) # Use heat kernel instead of adjacency matrix
            pseudo_label = self.p_net.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            one_hot_label = torch.argmax(pseudo_label, dim=1).squeeze() 
            output, dist_loss = saved_model.forward(feature, adjacency, eigenvalue, eigenvector, one_hot_label, is_train=False) 
 
            loss_test = F.nll_loss(output, label) + self.args.beta * dist_loss
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            ts = self.network.get_scales().cpu().detach()

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.eval()
        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.p_t) # Use heat kernel instead of adjacency matrix
            pseudo_label = self.p_net.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            pseudo_label = torch.argmax(pseudo_label, dim=1).squeeze()

            output, dist_loss = self.network.forward(feature, adjacency, eigenvalue, eigenvector, pseudo_label) 
 
            loss_test = F.nll_loss(output, label) + self.args.beta * dist_loss
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            ts = self.network.get_scales().cpu().detach()

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)


### Trainer for 'SVM'
class SVM_Trainer:
    def __init__(self, args, device, network, train_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.test_loader = test_loader
    
    def train(self):
        for adjacency, feature, label in self.train_loader:
            self.network.fit(feature, label)
    
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        for adjacency, feature, label in self.test_loader:
            output = self.network.predict(feature)
            output = torch.FloatTensor(encode_onehot(output))

            loss_test = torch.tensor([0])
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)

            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"One vs One - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None

class MLP_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
        
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                output = self.network.forward(feature) # Shape: (# of samples, # of labels)
                
                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")


                self.optimizer.step() # Updates the parameters

                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            output = self.network.forward(feature) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'GCN', 'GAT', 'GDC'
class GNN_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.optimizer = optimizer
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)

    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list = [] # List of val accuracy
        val_sens_list, val_prec_list = [], []

        best_val_acc = 0.7
        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                # if epoch % 100 == 0:
                print(f"\n Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()
            with torch.no_grad():
                for adjacency, feature, label in self.valid_loader:

                    output = self.network.forward(feature, adjacency) 

                    loss_val = F.nll_loss(output, label) 
                    accuracy_val = compute_accuracy(output, label)
                    val_acc_list.append(accuracy_val.item())
                    ac, pr, sp, se, f1 = confusion(output, label)

                    val_sens_list.append(se)
                    val_prec_list.append(pr)

                    if accuracy_val > best_val_acc:
                        best_val_acc = accuracy_val

                    print(' loss_val: {:.4f}'.format(loss_val.item()),
                          'acc_val: {:.4f}'.format(accuracy_val.item()), end='')
                    
                    wandb.log({"loss_val": loss_val.item(),
                          "accuracy_val": accuracy_val.item()})
        return val_acc_list, val_sens_list, val_prec_list

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label in self.test_loader:
            output = self.network.forward(feature, adjacency) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None


### Trainer for 'GraphHeat'
class GraphHeat_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_sz):
        self.args = args
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_sz = adj_sz
        self.optimizer = optimizer

        if args.use_t_local == 1: 
            self.t = torch.empty(adj_sz).fill_(2.)
        else:
            self.t = torch.tensor([2.])
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy

        for epoch in range(1, self.args.epochs + 1):
            self.network.train()

            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero
                
                heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label)
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                if epoch % 100 == 0:
                    print(f"Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}")

                self.optimizer.step() # Updates the parameters
                
                lt.append(loss_train.item())
                at.append(accuracy_train.item())

            self.network.eval()

    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s = [[] for _ in range(7)]

        self.network.eval()

        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel(eigenvalue, eigenvector, self.t)
            
            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label)
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), None
    

### Trainer for Exact
class Exact_Trainer:
    def __init__(self, args, device, network, optimizer, train_loader, valid_loader, test_loader, adj_size, cv_idx):
        self.args = args
        self.device = device
        self.network = network
        self.train_loader = train_loader
        self.valid_loader = valid_loader
        self.test_loader = test_loader
        self.adj_size = adj_size
        self.optimizer = optimizer
        self.cv_idx = cv_idx

        if args.use_t_local == 1: # Local scale
            self.t = torch.empty(adj_size).fill_(2.)
        else: # Global scale
            self.t = torch.tensor([2.]) 
        
        self.t_lr = self.args.t_lr
        self.t_loss_threshold = self.args.t_loss_threshold
        self.t_lambda = self.args.t_lambda
        self.t_threshold = self.args.t_threshold
        
        if torch.cuda.is_available():
            self.network = self.network.to(device)
            self.t = self.t.to(device)

    ### Scale regularization of loss function
    def t_loss(self):
        t_one = torch.abs(self.t)
        t_zero = torch.zeros_like(self.t)

        t_l = torch.where(self.t < self.t_loss_threshold, t_one, t_zero)

        return self.t_lambda * torch.sum(t_l)

    def t_deriv(self):
        t_one = self.t_lambda * torch.ones_like(self.t)
        t_zero = torch.zeros_like(self.t)

        t_de = torch.where(self.t < self.t_loss_threshold, -t_one, t_zero)

        return t_de

    def fir_deriv(self, output, feature, label, heat_kernel, heat_kernel_grad):
        y_oh = torch.zeros_like(output) # (# sample, # label)
        y_oh.scatter_(1, label.reshape(-1, 1), 1)
        dl_ds = (torch.exp(output) - y_oh) / output.shape[0]
        
        ds_dro0 = torch.mul(dl_ds, self.network.linrdp2) @ self.network.linear2.weight
        ds_dro1 = torch.mul(ds_dro0, self.network.linrdp)
        
        #ds_dro1 = torch.mul(dl_ds @ self.network.linear2.weight,  self.network.linrdp)
        dl_dro = torch.matmul(ds_dro1, self.network.linear.weight).reshape(-1, heat_kernel.shape[-2], self.args.hidden_units)
        
        if self.args.layer_num == 1:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp) @ self.network.gcn.gc1.weight.T
            dl_dt = torch.mul((dl_dc @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
        elif self.args.layer_num == 2:
            dl_dl2 = torch.mul(dl_dro, self.network.gcn.rdp2) @ self.network.gcn.gc2.weight.T

            dl_first = torch.mul((dl_dl2 @ self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad)
            backward = torch.matmul(self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2))

            dl_second_tmp = torch.mul(dl_dl2, self.network.gcn.rdp)
            dl_second = torch.matmul(torch.mul(dl_second_tmp @ backward, heat_kernel_grad), heat_kernel.swapaxes(1, 2))

            dl_dt = dl_first + dl_second
        elif self.args.layer_num == 3:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp3) @ self.network.gcn.gc3.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp2) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp) 
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third
        elif self.args.layer_num == 4:
            dl_dc = torch.mul(dl_dro, self.network.gcn.rdp4) @ self.network.gcn.gc4.weight.T
            dl_dc_first = torch.mul((dl_dc @ self.network.gcn.f3.swapaxes(1, 2)), heat_kernel_grad)
            dl_dc_second = torch.matmul(torch.mul(torch.matmul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.f2.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_third = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.f.swapaxes(1, 2)), heat_kernel_grad), 
                            heat_kernel.swapaxes(1, 2))
            dl_dc_fourth = torch.matmul(torch.mul(torch.matmul(torch.mul(torch.mul(torch.mul(dl_dc, self.network.gcn.rdp3) 
                            @ self.network.gcn.gc3.weight.T, self.network.gcn.rdp2)
                            @ self.network.gcn.gc2.weight.T, self.network.gcn.rdp)
                            @ self.network.gcn.gc1.weight.T, feature.swapaxes(1, 2)), heat_kernel_grad),
                            heat_kernel.swapaxes(1, 2))
            dl_dt = dl_dc_first + dl_dc_second + dl_dc_third + dl_dc_fourth

        if self.args.use_t_local == 1:
            dl_dt = torch.sum(dl_dt, dim=(0, 2))
        else:
            dl_dt = torch.tensor([torch.sum(dl_dt, dim=(0, 1, 2))]).to(self.device)
            
        dl_dt += self.t_deriv() # Add regularizer on t
        now_lr = self.t_lr * dl_dt

        now_lr[now_lr > self.t_threshold] = self.t_threshold
        now_lr[now_lr < -self.t_threshold] = -self.t_threshold

        self.t = self.t - now_lr # Update t

    ### Train
    def train(self):
        lt = [] # List of train loss
        at = [] # List of train accuracy
        val_acc_list = [] # List of val accuracy
        val_sens_list, val_prec_list = [], []
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="epoch"+str(0))
        for y in range(2, self.t.shape[0] + 2):
            ws.cell(row=1, column=y, value=2)
        i = 2

        best_val_acc = 0.8
        
        for epoch in range(1, self.args.epochs + 1):
            self.network.train()
            
            for adjacency, feature, label, eigenvalue, eigenvector in self.train_loader:
                self.optimizer.zero_grad() # Sets the gradients of all optimized tensors to zero

                heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                # Use heat kernel instead of adjacency matrix
                output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                loss_train = F.nll_loss(output, label) + self.t_loss()
                accuracy_train = compute_accuracy(output, label)
                loss_train.backward() # Computes the gradient of current tensor w.r.t. graph leaves
                
                # if epoch % 100 == 0:
                print(f"\n Epoch [{epoch} / {self.args.epochs}] loss_train: {loss_train.item():.5f} accuracy_train: {accuracy_train.item():.5f}", end="")

                with torch.no_grad():
                    self.fir_deriv(output, feature, label, heat_kernel, heat_kernel_grad)
                
                self.optimizer.step() # Updates the parameters
                
                wandb.log({"loss_train": loss_train.item(),
                          "acc_train": accuracy_train.item()})

                lt.append(loss_train.item())
                at.append(accuracy_train.item())
                
                s = self.t.detach().cpu().numpy()
                ws.cell(row=i, column=1, value="epoch"+str(epoch))
                for y in range(2, self.t.shape[0] + 2):
                    ws.cell(row=i, column=y, value=s[y-2])
                i += 1
    
            self.network.eval()
            with torch.no_grad():
                for adjacency, feature, label, eigenvalue, eigenvector in self.valid_loader:
                    heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t) # Use heat kernel instead of adjacency matrix

                    # Use heat kernel instead of adjacency matrix
                    output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)

                    loss_val = F.nll_loss(output, label) + self.t_loss()
                    accuracy_val = compute_accuracy(output, label)
                    val_acc_list.append(accuracy_val.item())
                    ac, pr, sp, se, f1 = confusion(output, label)

                    val_sens_list.append(se)
                    val_prec_list.append(pr)

                    if accuracy_val > best_val_acc:
                        if self.args.data == 'ppmi':
                            torch.save(self.network, 'trained_exact/ppmi/fold_' + str(self.cv_idx) + '_model.pt')
                            torch.save(self.t, 'trained_exact/ppmi/fold_' + str(self.cv_idx) + '_t.pt')
                        elif self.args.data == 'adni_fdg':
                            torch.save(self.network, 'trained_exact/adni_fdg/fold_' + str(self.cv_idx) + '_model.pt')
                            torch.save(self.t, 'trained_exact/adni_fdg/fold_' + str(self.cv_idx) + '_t.pt')
                        elif self.args.data == 'adni_ct':
                            torch.save(self.network, 'trained_exact/adni_ct/fold_' + str(self.cv_idx) + '_model.pt')
                            torch.save(self.t, 'trained_exact/adni_ct/fold_' + str(self.cv_idx) + '_t.pt')
                        print('saved!!')
                        best_val_acc = accuracy_val

                    print(' loss_val: {:.4f}'.format(loss_val.item()),
                          'acc_val: {:.4f}'.format(accuracy_val.item()), end='')
                    
                    wandb.log({"loss_val": loss_val.item(),
                          "accuracy_val": accuracy_val.item()})
        return val_acc_list, val_sens_list, val_prec_list

    ### Test
    def test(self):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        self.network.eval()

        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t)

            output = self.network.forward(feature, heat_kernel) # Shape: (# of samples, # of labels)
            
            loss_test = F.nll_loss(output, label) + self.t_loss()
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)
            print(f"Confusion - Accuracy: {ac:.10f} Precision: {pr:.10f} Specificity: {sp:.10f} Sensitivity: {se:.10f} F1 score: {f1:.10f}")

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            for i in range(len(self.t)):
                ts.append(self.t[i].item())

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)
    
    ### Test
    def load_and_test(self, saved_model, model_path):
        tl, ta, tac, tpr, tsp, tse, f1s, ts = [[] for _ in range(8)]

        saved_model = saved_model.to(self.device)
        saved_model.load_state_dict(torch.load(model_path))

        saved_model.eval()
        for adjacency, feature, label, eigenvalue, eigenvector in self.test_loader:
            heat_kernel, heat_kernel_grad = compute_heat_kernel_batch(eigenvalue, eigenvector, self.t)
           
            output = self.network.forward(feature, heat_kernel)
 
            loss_test = F.nll_loss(output, label) 
            accuracy_test = compute_accuracy(output, label)

            print("Prediction Labels >")
            print(output.max(1)[1])
            print("Real Labels >")
            print(label)
            
            print(f"Test set results: loss_test: {loss_test.item():.5f} accuracy_test: {accuracy_test.item():.5f}")

            ac, pr, sp, se, f1 = confusion(output, label)

            tl.append(loss_test.item())
            ta.append(accuracy_test.item())
            tac.append(ac.item())
            tpr.append(pr.item())
            tsp.append(sp.item())
            tse.append(se.item())
            f1s.append(f1.item())
            
            ts = self.network.get_scales().cpu().detach()

        return np.mean(tl), np.mean(ta), np.mean(tac), np.mean(tpr), np.mean(tsp), np.mean(tse), np.mean(f1s), np.array(ts)
